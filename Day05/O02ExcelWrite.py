
from openpyxl import load_workbook

wb = load_workbook("firstExcel.xlsx")

wb.create_sheet("Employees")

wb['Genpact'].active = False

wb.active = wb['Employees']

wb['Genpact'].active = False

print(wb.sheetnames)                # name of all the worksheets in the workbook

ws = wb.active

cell = ws["A4"]                     # current cell, add the data from next cell onwards

data = [
    ['empid', 'empname', 'age', 'dept', 'salary'],
    ['501', 'Jack', '23', 'HR', 35000],
    ['302', 'Jill', '29', 'Finance', 50000],
    ['403', 'Mike', '30', 'IT', 65000],
    ['604', 'Kevin', '35', 'MKT', 80000],
    ['805', 'Louis', '33', 'IT', 95000],
    ['906', 'Mesvin', '36', 'Accounts',60000],
]

for row in data:
    ws.append(row)

wb.save("firstExcel.xlsx")