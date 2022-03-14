
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart

wb = load_workbook('firstExcel.xlsx')

wb.create_sheet("Chart")

wb.active = wb['Chart']

ws = wb.active

data = [
    ('Product', 'Sales'),
    ('Pepsi', 450),
    ('Coke', 380),
    ('Thumps_up', 425),
    ('Sprite', 345),
    ('Maaza', 310),
    ('Tropicana', 485)
]

for row in data:
    ws.append(row)

dataRef = Reference(ws, min_row=2, min_col=2, max_row=7, max_col=2)
labelRef = Reference(ws, min_row=2, min_col=1, max_row=7)

chart = BarChart()
chart.add_data(dataRef)
chart.set_categories(labelRef)
chart.title = "Baverage Sales"
chart.x_axis.title = "Products"
chart.y_axis.title = "Sales in Lakhs"

ws.add_chart(chart, "E6")

wb.save('firstExcel.xlsx')