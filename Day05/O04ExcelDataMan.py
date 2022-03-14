
from openpyxl import load_workbook

wb = load_workbook("firstExcel.xlsx")

wb.active = wb['Employees']

ws = wb.active

for row in ws.iter_rows(min_row = 5, max_row=11, min_col=2):
    for cell in row:
        if cell.value == 'Kevin':
            cell.value = 'Kevin Costner'.upper()

wb.save("firstExcel.xlsx")