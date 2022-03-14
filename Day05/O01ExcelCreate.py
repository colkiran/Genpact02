
"""
file = Workbook
sheets(page) = Worksheet
rows    -   numbered - 1---1048576
columns -   named   - A - Z, AA - AZ, BA - BZ...XFD  = 16834
entering data into a cell
cell - intersection of row and col
cell - address - Col A, Row 10 - A10

"""
from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

ws = wb.active          # worksheet object

ws.title = "Genpact"

ws['C5'] = 2500
ws['C6'] = "Hello World"
ws['C7'] = datetime.now()

wb.save("firstExcel.xlsx")

