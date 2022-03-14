
from openpyxl import load_workbook

wb = load_workbook("firstExcel.xlsx")

wb.active = wb['Employees']

ws = wb.active

print(ws.dimensions)                # range of cell where data is present

for row in ws.iter_rows(min_row=5, min_col=1, max_row=11, max_col=5):
    for cell in row:
        print("{:^10s}".format(str(cell.value)), end=" ")
    print()
